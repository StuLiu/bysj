# 用于导出多个APP评论，目前效率较高
import time
import os
import threading
from openpyxl import Workbook
from commentDbHandler import CommentDbHandler
from dataBaseArgs import DataBaseArgs
from myLogFile import MyLogFile

class ThreadExportCommentOfOneApp(threading.Thread):

    def __init__(self, threadID, app, path, dbhandler):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.app = app
        self.path = path
        self.dbhandler = dbhandler


    def run(self):
        # get comments from db and save them to app_id-app_name.xlsx
        print('query %s\'s comments......' % (self.app[1]))
        lock.acquire()
        comments = self.dbhandler.queryCommentsByAppId(self.app[0])
        lock.release()
        # print('run', comments)
        xlsxFileName = self.path + "/" + self.app[0] + "_" + self.app[1] + ".xlsx"
        wb = Workbook()
        ws = wb.active
        # app_id,time,comment_id,title,content,voteSum,voteCount,rating,version,user_name,isSpam,app_name
        head = ('app_id', 'time', 'comment_id', 'title', 'content', 'voteSum', 'voteCount',
                'rating', 'version', 'user_name', 'isSpam', 'app_name')
        for j in range(len(head)):
            ws.cell(row=1, column=j + 1).value = head[j]
        for comment in comments:
            ws.append(comment)
        wb.save(filename=xlsxFileName)
        print('save %s finished,total %d comments.\n' % (self.app[1], len(comments)))

class DbToXlsx(object):

    def __init__(self,label=''):
        self.label = label
        # connet to database
        self.dbHandler = CommentDbHandler('127.0.0.1', DataBaseArgs.getUserName(),
                                     DataBaseArgs.getPassWord(), DataBaseArgs.getDataBase())

    def exportAllComments(self):
        currTime = time.localtime(time.time())
        xlsxFileName = "./output/"+ str(currTime.tm_year) + "_" + str(currTime.tm_mon) + \
                       "_" + str(currTime.tm_mday) + "_" + str(currTime.tm_hour) +\
                       "_" + str(currTime.tm_min) + "_" + str(currTime.tm_sec) + '.xlsx'

        comments = self.dbHandler.queryComments()

        wb = Workbook()
        ws = wb.active
        # app_id,time,comment_id,title,content,voteSum,voteCount,rating,version,user_name,isSpam,app_name
        head = ('app_id', 'time', 'comment_id', 'title', 'content', 'voteSum', 'voteCount',
                'rating', 'version', 'user_name', 'isSpam', 'app_name')
        for j in range(len(head)):
            ws.cell(row=1, column=j + 1).value = head[j]
        for comment in comments:
            ws.append(comment)
        wb.save(filename=xlsxFileName)
        global logTexts
        logTexts.append('save all comments to '+xlsxFileName+' finished.')

    # export comments to several xlsx files named by app_id-app_name
    # multi thread
    def exportCommentsEachApp(self):
        # create output dir
        currTime = time.localtime(time.time())  # get current time
        dirPath = "./output/" + str(currTime.tm_year) + "_" + str(currTime.tm_mon) + "_" +\
                  str(currTime.tm_mday) + "_" + str(currTime.tm_hour) + "_" + str(currTime.tm_min) +\
                  "_" + str(currTime.tm_sec) + 'xlsx'
        os.mkdir(dirPath)
        appTuple = self.dbHandler.getApps()  # ((app_id,app_name),(app_id,app_name),...)

        threads = []
        for i in range(len(appTuple)):
            print(self.dbHandler.getApps()[i])
            myThread = ThreadExportCommentOfOneApp(i,appTuple[i],dirPath,self.dbHandler)
            threads.append(myThread)
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()
        global logTexts
        logTexts.append('save comments to app_id-app_name.xlsx files finished.(_2)')


if __name__ == '__main__':
    lock = threading.Lock()
    logTexts = []
    obj = DbToXlsx()
    try:
        fr = time.time()
        obj.exportCommentsEachApp()
        to = time.time()
        logTexts.append('runtime:' + str(to - fr) + 's')
        print(logTexts)
        MyLogFile('Log').addMsg(logTexts)
    except Exception as e:
        print(e)
        MyLogFile('Log').addMsg([str(e)])
        exit(1)