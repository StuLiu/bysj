
"""
将xlsx文件中已标注评论转移到数据库中
note：一个xlsx文件对应一个app的评论
"""

print(__doc__)

from openpyxl.reader.excel import load_workbook
from commentDbHandler import CommentDbHandler
from dataBaseArgs import DataBaseArgs

class XlsxToDb(object):

    def __init__(self):
        self.dbHandler = CommentDbHandler('127.0.0.1', DataBaseArgs.getUserName(),
                                          DataBaseArgs.getPassWord(), DataBaseArgs.getDataBase())

    # 将input目录下某个xlsx文件的评论导入到数据库
    def executeOneApp(self, fileName):
        print(fileName)
        wb = None
        try:
            wb = load_workbook(filename='signedComments/'+fileName+'.xlsx')
            ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        except Exception:
            print("未找到"+fileName+".xlsx文件")
            return
        appId = ws.cell(row=2, column=1).value
        # # 显示有多少张表
        # print( "Worksheet range(s):", wb.get_named_ranges() )
        # print( "Worksheet name(s):", wb.get_sheet_names() )
        # # 显示表名，表行数，表列数
        # print( "Work Sheet Titile:", ws.title )
        # print( "Work Sheet Rows:", ws.max_row)
        # print( "Work Sheet Cols:", ws.max_column )

        # 建立存储数据的列表
        comments_list = []
        # 把数据存到字典中

        for row in range(2, ws.max_row + 1):
            temp_list = []
            for col in range(2,12):
                temp_list.append( ws.cell(row=row, column=col).value )
            temp_list.insert(9,appId)
            self.dbHandler.insertSignedComments(temp_list)
            comments_list.append(temp_list)

        # 打印字典数据个数
        print('Total:%d' % len(comments_list))

    # 将input目录下所有已标记xlsx文件的评论导入到数据  库
    def executeAllApp(self):
        appList = list(self.dbHandler.getApps())
        print(len(appList))
        for app in appList:
            self.executeOneApp(app[0]+"_"+app[1])

if __name__ == "__main__":
    xlsxToDb = XlsxToDb()
    xlsxToDb.executeAllApp()
