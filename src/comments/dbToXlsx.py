# 用于导出多个APP评论，目前效率较高
import time
import os
import threading
import config
from openpyxl import Workbook
from database.app_comments_dbHandler import AppCommentsDbHandler
from database.apple_app_dbHandler import AppleAppDbHandler

class ThreadExportCommentOfOneApp(threading.Thread):

    def __init__(self, threadID, app, path, dbhandler):
        threading.Thread.__init__(self)
        self.__threadID = threadID
        self.__app = app
        self.__path = path
        self.__dbhandler = dbhandler


    def run(self):
        # get comments from db and save them to app_id-app_name.xlsx
        print('query %s\'s comments......' % (self.__app[1]))
        lock.acquire()
        comments = self.__dbhandler.queryCommentsByAppId(self.__app[0])
        lock.release()
        # print('run', comments)
        xlsxFileName = os.path.join(self.__path, self.__app[0] + "_" + self.__app[1] + ".xlsx")
        wb = Workbook()
        ws = wb.active
        # app_id,time,comment_id,title,content,voteSum,voteCount,rating,version,user_name,isSpam,app_name
        head = ('app_id', 'time', 'comment_id', 'title', 'content', 'voteSum', 'voteCount',
                'rating', 'version', 'user_name', 'isSpam', 'app_name')
        for j in range(len(head)):
            ws.cell(row = 1, column = j + 1).value = head[j]
        for comment in comments:
            ws.append(comment)
        wb.save(filename = xlsxFileName)
        print('save %s finished,total %d comments.\n' % (self.__app[1], len(comments)))

class DbToXlsx(object):

    def __init__(self):
        # connet to database
        self.__appCommentsHandler = AppCommentsDbHandler()
        self.__appleAppHandler = AppleAppDbHandler()

    def exportAllComments(self):
        currTime = time.localtime(time.time())
        xlsxFileName = str(currTime.tm_year) + "_" + str(currTime.tm_mon) + \
                       "_" + str(currTime.tm_mday) + "_" + str(currTime.tm_hour) +\
                       "_" + str(currTime.tm_min) + "_" + str(currTime.tm_sec) + '.xlsx'
        filePath = os.path.join(config.RESOURCES_PATH,'output',xlsxFileName)
        comments = self.__appCommentsHandler.queryAll()

        wb = Workbook()
        ws = wb.active
        # app_id,time,comment_id,title,content,voteSum,voteCount,rating,version,user_name,isSpam,app_name
        head = ('app_id', 'time', 'comment_id', 'title', 'content', 'voteSum', 'voteCount',
                'rating', 'version', 'user_name', 'isSpam', 'app_name')
        for j in range(len(head)):
            ws.cell(row=1, column=j + 1).value = head[j]
        for comment in comments:
            ws.append(comment)
        wb.save(filename = filePath)

    # export comments to several xlsx files named by app_id-app_name
    # multi thread
    def exportCommentsEachApp(self):
        # create output dir
        currTime = time.localtime(time.time())  # get current time
        dirName = str(currTime.tm_year) + "_" + str(currTime.tm_mon) + "_" +\
                  str(currTime.tm_mday) + "_" + str(currTime.tm_hour) + "_" + str(currTime.tm_min) +\
                  "_" + str(currTime.tm_sec) + 'xlsx'
        dirPath = os.path.join(config.RESOURCES_PATH, 'output', dirName)
        os.mkdir(dirPath)
        appTuple = self.__appleAppHandler.queryAll()  # ((app_id,app_name),(app_id,app_name),...)

        threads = []
        for i in range(len(appTuple)):
            print(appTuple[i])
            myThread = ThreadExportCommentOfOneApp(i,appTuple[i],dirPath,self.__appCommentsHandler)
            threads.append(myThread)
        for thread in threads:
            thread.start()
        for thread in threads:
            thread.join()


if __name__ == '__main__':
    lock = threading.Lock()
    try:
        fr = time.time()
        # DbToXlsx().exportAllComments()
        DbToXlsx().exportCommentsEachApp()
        to = time.time()
    except Exception as e:
        print(e)
        exit(1)